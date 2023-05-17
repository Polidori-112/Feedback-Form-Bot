import openpyxl
import os, os.path
import pandas as pd


#used to take last name from name string (protect against faulty inputs)
def name(s):
  # Split the string into a list of words
  try:
    words = s.split()
  except:
    return s
  
  # Return the last word in the list in lowercase
  return words[-1].lower()



#used to capitalize the first letter of a name when outputting to txt
def name_out(s):
  # Get the first character of the string
  first_char = s[0]

  # Capitalize the first character
  capitalized_char = first_char.upper()

  # Return a new string with the capitalized first character and the rest of the original string
  return capitalized_char + s[1:]



#remove the prompt questions from the spreadsheet
def remove_prompt(file_name):
  # Load the workbook
  wb = openpyxl.load_workbook(file_name)

  # Get the active sheet
  sheet = wb.active
  if (sheet["A1"].value == 'Timestamp'):
    # Delete the first row
    sheet.delete_rows(1, 1)

    # Save the changes to the workbook
    wb.save(file_name)

#Checks if the value of the cell read in is empty or not
def emptyCell(s):
  if (s):
    return 1
  else:
    return 0
    
  try:
    first_word = s.split()[0]  # get the first word of the string
    if first_word == 'Unnamed':
        return 1
    else:
        return 0
  except:
    return 1

#get the doing well cell
def well_cell(cell, file_name):
  # Load the workbook
  wb = openpyxl.load_workbook(file_name)

  # Get the active sheet
  sheet = wb.active
  
  # Get the column index of the cell
  col_index = cell.column

  # Calculate the column index of the cell to the right
  right_col_index = col_index + 1

  # Convert the column index to a column letter
  right_col_letter = openpyxl.utils.get_column_letter(right_col_index)

  # Return the cell to the right
  return sheet[f"{right_col_letter}{cell.row}"]
  


#get the improvement cell
def impr_cell(cell, file_name):
  # Load the workbook
  wb = openpyxl.load_workbook(file_name)

  # Get the active sheet
  sheet = wb.active
  
  # Get the column index of the cell
  col_index = cell.column

  # Calculate the column index of the cell to the right
  right_col_index = col_index + 2

  # Convert the column index to a column letter
  right_col_letter = openpyxl.utils.get_column_letter(right_col_index)

  # Return the cell to the right
  return sheet[f"{right_col_letter}{cell.row}"]
  
  
  
#get the other cell
def other_cell(cell, file_name):
  # Load the workbook
  wb = openpyxl.load_workbook(file_name)

  # Get the active sheet
  sheet = wb.active
  
  # Get the column index of the cell
  col_index = cell.column

  # Calculate the column index of the cell to the right
  right_col_index = col_index + 3

  # Convert the column index to a column letter
  right_col_letter = openpyxl.utils.get_column_letter(right_col_index)

  # Return the cell to the right
  return sheet[f"{right_col_letter}{cell.row}"]



#generate/append well cell info into the mid's .txt document
def append_well(column, file_name):
  # Load the workbook
  wb = openpyxl.load_workbook(file_name)

  # Get the active sheet
  sheet = wb.active

  # Iterate over the cells in the specified column
  for cell in sheet[column]:
    # Get the cell value
    midn = name(cell.value)
    feedback = well_cell(cell, file_name).value
    
    if feedback:
      # Create a new text file with the cell value as the contents
      with open(f"emails/{midn}.txt", "a") as f:
        f.write(feedback + "\n")
      


#generate/append improvement info into the mid's .txt document
def append_impr(column, file_name):
  # Load the workbook
  wb = openpyxl.load_workbook(file_name)

  # Get the active sheet
  sheet = wb.active

  # Iterate over the cells in the specified column
  for cell in sheet[column]:
    # Get the cell value
    midn = name(cell.value)
    feedback = impr_cell(cell, file_name).value
    
    if feedback:
      # Create a new text file with the cell value as the contents
      with open(f"emails/{midn}.txt", "a") as f:
        f.write(feedback + "\n")
      
      

#generate/append improvement info into the mid's .txt document
def append_other(column, file_name):
  # Load the workbook
  wb = openpyxl.load_workbook(file_name)

  # Get the active sheet
  sheet = wb.active

  # Iterate over the cells in the specified column
  for cell in sheet[column]:
    # Get the cell value
    midn = name(cell.value)
    feedback = other_cell(cell, file_name).value

    if (feedback):
      # Create a new text file with the cell value as the contents
      with open(f"emails/{midn}.txt", "a") as f:
        f.write(feedback + "\n")

#generate/append exceptional cell info into the mid's .txt document
def append_exceptional(column, file_name):
  # Load the workbook
  wb = openpyxl.load_workbook(file_name)

  # Get the active sheet
  sheet = wb.active

  # Iterate over the cells in the specified column
  for cell in sheet[column]:
    # Get the cell value
    midn = name(cell.value)
    feedback = well_cell(cell, file_name).value

    if feedback:      
      # Create a new text file with the cell value as the contents
      with open(f"emails/{midn}.txt", "a") as f:
        f.write(f"\nAdditionally, you were noted for being an excellent MIDN for:\n{feedback}\n")


#generate mid's .txt document
def generate_txt(file_name):
  column = "B"
  # Load the workbook
  wb = openpyxl.load_workbook(file_name)

  # Get the active sheet
  sheet = wb.active

  # Iterate over the cells in the specified column
  for cell in sheet[column]:
    # Get the cell value
    midn = name(cell.value)

    # Create a new text file with the cell value as the contents
    if midn:    
      f = open(f"emails/{midn}.txt", "a")
      f.write(f"Good Morning MIDN {name_out(midn)},\n\nBelow are the results from your peer feedback.\n\nQuestion: What does this MIDN do well?\nResponses:\n\n")


#appends information to all .txt files
def append_all(directory, str):
  # Iterate over the files in the directory
  for file_name in os.listdir(directory):
    # Check if the file is a .txt file
    if file_name.endswith(".txt"):
      # Open the file in append mode
      with open(os.path.join(directory, file_name), "a") as f:
        # Write the word to the file
        f.write(str)



#writes all of the information in the spreadsheet to .txt files
def write_txt(file):
  #create the intro blurb
  generate_txt(file)
  
  #add the well cells
  for letter in ["D", "H", "L", "P", "T", "X", "AB", "AD", "AF"]:
    append_well(letter, file)
    
  print("Progress: 25%")
  #generate improve blurb
  str = "\n\nQuestion: What does this MIDN need to improve on?\nResponses:\n\n"

  append_all("emails/", str)

  #add the improve cells
  for letter in ["D", "H", "L", "P", "T", "X", "AB", "AD", "AF"]:
    append_impr(letter, file)
  
  print("Progress: 50%")
  #generate other blurb
  str = "\n\nQuestion: Other feedback for this MIDN?\nResponses:\n\n"
  append_all("emails/", str)
  
  print("Progress: 75%\n")
  #add the other cells
  for letter in ["D", "H", "L", "P", "T", "X", "AB", "AD", "AF"]:
    append_other(letter, file)

  #add the exceptional cells
  for letter in ["AJ", "AL"]:
    append_exceptional(letter, file)

  #generate final blurb
  str = "\nVery Respectfully,\nS3-LD Bot"

  append_all("emails/", str)
  

#delete all the old .txt files in the email directory
def deletePrevData():
    #remove all files in emails directory
    dir_path = "emails"
    
    # Get a list of all files in the directory
    files = os.listdir(dir_path)
    
    # Loop through the files and remove them one by one
    for email in files:
        file_path = os.path.join(dir_path, email)
        if os.path.isfile(file_path):
            os.remove(file_path)

def removeUnwantedCharacters(file):
    # Load the spreadsheet into a Pandas DataFrame
    df = pd.read_excel(file)
    
    # Remove the '/' character from every cell in the DataFrame
    # add in any characters people entered here that caused the program to crash
    df = df.replace('/', ' ', regex=True)
    
    # Save the modified DataFrame back to the spreadsheet
    df.to_excel(file, index=False)

def jank_fix(directory):
    for filename in os.listdir(directory):
        if filename.endswith('.txt'): # or whatever file extension you're working with
            filepath = os.path.join(directory, filename)
            with open(filepath, 'r') as f:
                lines = f.readlines()
            with open(filepath, 'w') as f:
                for line in lines:
                    if not line.startswith('Unnamed'):
                        f.write(line)
        if filename[0].isdigit():
           os.remove(os.path.join(directory, filename))

#main code
file = "spreadsheets/Feedback.xlsx"


#generate error code if applicable
try:
  #clear the emails directory
  deletePrevData()
  #remove characters that would cause the program to crash
  removeUnwantedCharacters(file)
  #remove the first line of the spreadsheet if it is not desired data
  remove_prompt(file)
  print("Writing files...\n(This may take a minute)")
  #run the actual program
  write_txt(file)
  #fix this weird bug where some empty cells are read as 'unnamed' instead of null
  jank_fix("emails")
except:
  print("An error has occurred. Make sure that the responses are in a file titled 'Feedback.xlsx' in the spreadsheets folder (.../S3LDBot/spreadsheets/Feedback.xlsx). If that is the case, someone probably had faulty input to the form. Try removing the try-except statement towards the end to debug it or ask a CS friend to help out")
  exit(0)
    
#final statements
print("All .txt files written with no errors")
print("\nIMPORTANT: PLEASE READ THE README.txt DOCUMENT BEFORE CONTINUING")
print("\nPlease ensure .txt files are correct, then run gmail.py to send out information.\nThank you for choosing S3-LD Bot.")
